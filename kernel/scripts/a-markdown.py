#!/usr/bin/env python3
"""Convierte un insumo binario a markdown para que un LLM lo lea sin quemar tokens.

Determinista y sin modelo: el binario nunca entra en contexto. Se invoca desde
/x-procesar-inbox (paso 4) y deja un .md que cita al original crudo.

    python3 a-markdown.py <archivo> [--out DIR] [--filas N] [--origen RUTA] [--stdout]

Motor por formato y el porqué de cada elección: ver README.md de esta carpeta.
"""

import base64
import datetime
import html
import os
import re
import sys
import urllib.parse
import zlib
import xml.etree.ElementTree as ET
from pathlib import Path

_WIN = os.name == "nt"

# La consola de Windows no usa UTF-8 por defecto: sin esto, imprimir «—», «…» o
# cualquier acento en --stdout revienta con UnicodeEncodeError.
for _flujo in (sys.stdout, sys.stderr):
    if hasattr(_flujo, "reconfigure"):
        _flujo.reconfigure(encoding="utf-8")

# El venv hermano trae markitdown y openpyxl. Re-ejecutarse dentro permite
# invocar el script con el python del sistema sin activar nada.
_VENV = Path(__file__).resolve().parent / ".venv"
_VENV_PY = _VENV / ("Scripts" if _WIN else "bin") / ("python.exe" if _WIN else "python3")

# Se compara sys.prefix, no el ejecutable: en POSIX .venv/bin/python3 es un symlink
# al python real, así que resolve() en ambos lados daría el mismo binario y nunca
# se re-ejecutaría. La guarda de entorno evita un bucle si algo falla.
if _VENV_PY.exists() and Path(sys.prefix) != _VENV and not os.environ.get("_A_MARKDOWN_VENV"):
    os.environ["_A_MARKDOWN_VENV"] = "1"
    if _WIN:
        # os.execv en Windows no reemplaza el proceso: lanza otro y mata el actual,
        # con lo que la shell da por terminado el comando antes de tiempo.
        import subprocess

        sys.exit(subprocess.run([str(_VENV_PY), *sys.argv]).returncode)
    os.execv(str(_VENV_PY), [str(_VENV_PY), *sys.argv])

# markitdown exige >=3.10 y ningún sistema lo garantiza: macOS trae 3.9, Windows no
# trae nada, y las distros varían. Sin esta guarda el fallo aparece más tarde y más
# lejos (un error de pip, o un "falta markitdown" con markitdown instalado).
if sys.version_info < (3, 10):
    _actual = ".".join(map(str, sys.version_info[:3]))
    _comos = {
        "darwin": "El Python que trae macOS es 3.9 y se queda corto. Instalar uno propio:\n"
                  "    brew install python@3.13",
        "win32": "Instalar Python desde python.org, o:\n"
                 "    winget install Python.Python.3.13",
    }
    _como = _comos.get(sys.platform, "Instalar Python 3.10 o superior. En Debian/Ubuntu:\n"
                                     "    sudo apt install python3.12 python3.12-venv")
    sys.exit(f"ERROR: a-markdown.py necesita Python 3.10+ y este es {_actual}.\n\n  {_como}\n\n"
             "  Después, recrear el venv con ese intérprete:\n"
             "    rm -rf kernel/scripts/.venv   (en Windows: rmdir /s kernel\\scripts\\.venv)\n"
             "  y volver a seguir kernel/scripts/README.md")

FILAS_POR_DEFECTO = 50
LEGACY = {".doc": ".docx", ".ppt": ".pptx", ".xls": ".xlsx"}


class ErrorConversion(Exception):
    """Fallo esperado y accionable — se reporta sin traceback."""


# ---------------------------------------------------------------- utilidades

def _leer(ruta, errores="strict"):
    """Lee texto descartando el BOM que meten Notepad y varios exportadores de Windows.

    utf-8-sig lo quita si está y se comporta como utf-8 si no. ElementTree tolera el
    BOM, pero en el passthrough de .yaml se colaría como un ﻿ invisible al frente
    del bloque de código, que parece corrupción y no aporta nada.
    """
    return ruta.read_text(encoding="utf-8-sig", errors=errores)


def _miles(n):
    return f"{n:,}".replace(",", ".")


def _celda(v):
    """Un valor de celda/tabla como texto seguro dentro de una tabla markdown."""
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:19]
    return re.sub(r"\s+", " ", str(v)).replace("|", r"\|").strip()


def _tabla_md(filas):
    """filas: lista de tuplas. La primera es la cabecera. Recorta columnas vacías."""
    ancho = 0
    for f in filas:
        for i, c in enumerate(f):
            if _celda(c):
                ancho = max(ancho, i + 1)
    if not ancho:
        return []
    out = []
    for n, f in enumerate(filas):
        celdas = [_celda(c) for c in list(f)[:ancho]]
        celdas += [""] * (ancho - len(celdas))
        out.append("| " + " | ".join(celdas) + " |")
        if n == 0:
            out.append("|" + "---|" * ancho)
    return out


def _limpiar_etiqueta(s):
    """Las etiquetas de drawio vienen con HTML incrustado."""
    s = re.sub(r"<br\s*/?>", " ", s or "", flags=re.I)
    s = re.sub(r"<[^>]+>", "", s)
    s = html.unescape(s).replace("\xa0", " ")
    return re.sub(r"\s+", " ", s).strip()


# ---------------------------------------------------------------- motores

def _markitdown(ruta):
    """docx / html / pdf — aquí MarkItDown es exactamente la herramienta correcta."""
    try:
        from markitdown import MarkItDown
    except ImportError as e:
        py = "python" if _WIN else "python3"
        pip = _VENV / ("Scripts" if _WIN else "bin") / "pip"
        raise ErrorConversion(
            f"Falta markitdown. Instalar:\n"
            f"  {py} -m venv kernel/scripts/.venv\n"
            f"  {pip} install -r kernel/scripts/requirements.txt"
        ) from e
    return (MarkItDown().convert(str(ruta)).text_content or "").strip()


def conv_docx(ruta, _filas):
    return _markitdown(ruta), [], "Insumo"


# Chatarra de plantilla web: en una página real son cientos de tokens de ruido
# que MarkItDown por sí solo no quita (sí quita script/style, pero no esto).
RUIDO_HTML = ("nav", "footer", "header", "aside", "form", "script", "style", "noscript", "svg")


def conv_html(ruta, _filas):
    import tempfile

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return _markitdown(ruta), ["Sin bs4: no se pudo limpiar la plantilla (nav/footer)."], "Insumo"

    sopa = BeautifulSoup(_leer(ruta, errores="replace"), "html.parser")
    quitados = 0
    for tag in sopa.find_all(RUIDO_HTML):
        tag.decompose()
        quitados += 1
    # El contenido real, si la página lo marca; si no, todo lo que quedó.
    principal = sopa.find("main") or sopa.find("article") or sopa
    with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False, encoding="utf-8") as tmp:
        tmp.write(str(principal))
        temporal = Path(tmp.name)
    try:
        texto = _markitdown(temporal)
    finally:
        temporal.unlink(missing_ok=True)
    avisos = [f"Se descartaron {quitados} bloques de plantilla (nav/footer/script…)."] if quitados else []
    return texto, avisos, "Insumo"


def conv_pdf(ruta, _filas):
    texto = _markitdown(ruta)
    avisos = []
    try:
        from pdfminer.pdfpage import PDFPage

        with open(ruta, "rb") as f:
            paginas = len(list(PDFPage.get_pages(f)))
    except Exception:
        paginas = 0
    # Un PDF escaneado extrae casi nada: hay que decirlo, no devolver vacío en silencio.
    if paginas and len(texto) / paginas < 100:
        avisos.append(
            f"PDF probablemente escaneado o sin capa de texto: {paginas} páginas y solo "
            f"{_miles(len(texto))} caracteres extraídos. Lo convertido puede estar incompleto; "
            f"el original en raw/ sigue siendo la fuente."
        )
    return texto, avisos, "Insumo"


def conv_pptx(ruta, _filas):
    """MarkItDown ya emite diapositivas, tablas Y notas del orador (bajo `### Notes:`).

    Aquí hubo un conversor propio con python-pptx, escrito sobre la creencia — repetida
    en varios artículos — de que MarkItDown se saltaba las notas del orador. Se midió
    contra markitdown 0.1.6 y es falso: las incluye, y con mejor estructura. El código
    propio se borró. Si algún día se comprueba que se pierden, aquí va el reemplazo.
    """
    return _markitdown(ruta), [], "Insumo"


def _formulas_sin_cache(ruta, tope=200):
    """¿Hay fórmulas sin valor calculado en caché?

    openpyxl no evalúa fórmulas: lee el valor que Excel dejó guardado. Un libro
    generado por un script y nunca abierto en Excel no tiene ese caché, así que
    las columnas calculadas saldrían vacías SIN avisar. Hay que decirlo.
    """
    from openpyxl import load_workbook

    wf = load_workbook(str(ruta), data_only=False, read_only=True)
    wv = load_workbook(str(ruta), data_only=True, read_only=True)
    try:
        for hf, hv in zip(wf.worksheets, wv.worksheets):
            for n, (ff, fv) in enumerate(zip(hf.iter_rows(values_only=True),
                                             hv.iter_rows(values_only=True))):
                if n > tope:
                    break  # si hay fórmulas en una columna, están arriba
                for cf, cv in zip(ff, fv):
                    if isinstance(cf, str) and cf.startswith("=") and cv is None:
                        return True
        return False
    finally:
        wf.close()
        wv.close()


def conv_xlsx(ruta, filas):
    """La bomba de tokens nº1: se muestrea. El .xlsx íntegro queda en raw/."""
    from openpyxl import load_workbook

    wb = load_workbook(str(ruta), data_only=True, read_only=True)
    partes, avisos = [], []
    if _formulas_sin_cache(ruta):
        avisos.append(
            "El libro tiene fórmulas sin valor calculado en caché: las columnas calculadas "
            "salen VACÍAS. Suele pasar con libros generados por script y nunca abiertos en "
            "Excel. Para materializarlas: abrir en Excel/LibreOffice y guardar."
        )

    for ws in wb.worksheets:
        conservadas, total, ancho = [], 0, 0
        for fila in ws.iter_rows(values_only=True):
            if fila is None or not any(_celda(c) for c in fila):
                continue
            total += 1
            ancho = max(ancho, max(i + 1 for i, c in enumerate(fila) if _celda(c)))
            if len(conservadas) < filas + 1:
                conservadas.append(fila)

        if not total:
            continue

        partes.append(f"### Hoja: {ws.title}")
        partes.append("")
        if total > len(conservadas):
            nota = (f"> {_miles(total)} filas × {ancho} columnas — se muestran las primeras "
                    f"{_miles(len(conservadas) - 1)}. El original completo está en raw/.")
            avisos.append(f"Hoja '{ws.title}': truncada a {_miles(len(conservadas) - 1)} "
                          f"de {_miles(total)} filas.")
        else:
            nota = f"> {_miles(total)} filas × {ancho} columnas (completa)."
        partes += [nota, ""] + _tabla_md(conservadas) + [""]

    wb.close()
    if not partes:
        avisos.append("El libro no tiene ninguna hoja con datos.")
    return "\n".join(partes).strip(), avisos, "Insumo"


def conv_drawio(ruta, _filas):
    """Ninguna herramienta del mercado lo cubre. Encaja con la regla: diagramas como texto."""
    raiz = ET.fromstring(_leer(ruta))
    diagramas = raiz.findall(".//diagram") or [raiz]
    partes, avisos = [], []

    for d in diagramas:
        modelo = d.find(".//mxGraphModel")
        if modelo is None:
            # Variante comprimida: base64 + deflate crudo + url-encode.
            crudo = (d.text or "").strip()
            if not crudo:
                continue
            try:
                xml = zlib.decompress(base64.b64decode(crudo), -15).decode("utf-8")
                modelo = ET.fromstring(urllib.parse.unquote(xml))
            except Exception as e:
                raise ErrorConversion(
                    f"No se pudo descomprimir la página '{d.get('name') or '?'}' de {ruta.name}: {e}"
                ) from e

        nodos, aristas = {}, []
        envueltos = {id(mc) for o in modelo.iter()
                     if o.tag in ("object", "UserObject") for mc in o.findall("mxCell")}

        for el in modelo.iter():
            if el.tag in ("object", "UserObject"):
                mc = el.find("mxCell")
                if mc is None:
                    continue
                cid, etiqueta, attrs = el.get("id"), el.get("label") or "", mc
            elif el.tag == "mxCell" and id(el) not in envueltos:
                cid, etiqueta, attrs = el.get("id"), el.get("value") or "", el
            else:
                continue
            if attrs.get("vertex") == "1":
                nodos[cid] = _limpiar_etiqueta(etiqueta)
            elif attrs.get("edge") == "1":
                aristas.append((attrs.get("source"), attrs.get("target"),
                                _limpiar_etiqueta(etiqueta)))

        referenciados = {s for s, _, _ in aristas} | {t for _, t, _ in aristas}
        visibles = {c: e for c, e in nodos.items() if e or c in referenciados}
        if not visibles:
            continue

        ids = {c: f"n{i}" for i, c in enumerate(visibles)}
        lineas = ["flowchart TD"]
        for c, etiqueta in visibles.items():
            texto = (etiqueta or "(sin etiqueta)").replace('"', "#quot;")
            lineas.append(f'    {ids[c]}["{texto}"]')
        for origen, destino, etiqueta in aristas:
            if origen not in ids or destino not in ids:
                continue  # arista suelta: sin extremo no dice nada
            if etiqueta:
                lineas.append(f'    {ids[origen]} -->|"{etiqueta.replace(chr(34), "#quot;")}"| {ids[destino]}')
            else:
                lineas.append(f"    {ids[origen]} --> {ids[destino]}")

        nombre = d.get("name")
        if len(diagramas) > 1 and nombre:
            partes.append(f"## {nombre}")
            partes.append("")
        partes += ["```mermaid"] + lineas + ["```", ""]
        n_nodos, n_aristas = len(visibles), len(aristas)
        avisos.append(f"Página '{nombre or ruta.stem}': "
                      f"{n_nodos} nodo{'s' if n_nodos != 1 else ''}, "
                      f"{n_aristas} conexi{'ones' if n_aristas != 1 else 'ón'}.")

    if not partes:
        raise ErrorConversion(f"{ruta.name} no contiene ningún diagrama con nodos.")
    return "\n".join(partes).strip(), avisos, "Diagrama"


def conv_yaml(ruta, _filas):
    """Ya es texto casi óptimo para un LLM: convertirlo costaría tokens y perdería fidelidad."""
    return "```yaml\n" + _leer(ruta).strip() + "\n```", [], "Insumo"


MOTORES = {
    ".docx": conv_docx, ".pdf": conv_pdf, ".html": conv_html, ".htm": conv_html,
    ".pptx": conv_pptx, ".xlsx": conv_xlsx, ".drawio": conv_drawio,
    ".yaml": conv_yaml, ".yml": conv_yaml,
}


# ---------------------------------------------------------------- salida

def _raiz_bundle(ruta):
    for p in [ruta.resolve()] + list(ruta.resolve().parents):
        if (p / "kernel").is_dir() and (p / "cerebro").is_dir():
            return p
    return None


def _origen_bundle(ruta, origen):
    if origen:
        return origen
    raiz = _raiz_bundle(ruta)
    if raiz:
        try:
            return "/" + str(ruta.resolve().relative_to(raiz))
        except ValueError:
            pass
    return ruta.name


def construir(ruta, filas, origen):
    ext = ruta.suffix.lower()
    if ext in LEGACY:
        raise ErrorConversion(
            f"{ruta.name} está en formato binario legacy ({ext}), que no se puede leer.\n"
            f"Abrilo en Office y usá «Guardar como» → {LEGACY[ext]}, y volvé a dejarlo en inbox/."
        )
    if ext not in MOTORES:
        raise ErrorConversion(
            f"No hay motor para {ext}. Soportados: {', '.join(sorted(MOTORES))}."
        )
    if not ruta.is_file():
        raise ErrorConversion(f"No existe el archivo: {ruta}")

    cuerpo, avisos, tipo = MOTORES[ext](ruta, filas)
    if not cuerpo.strip():
        avisos.append("La conversión no extrajo texto. Revisar el original en raw/.")

    # Se quitan el prefijo de fecha y el sufijo tiny_uuid del archivado en raw/
    # (YYYY-MM-DD-descripcion-tiny_uuid.ext); el uuid debe traer algún dígito para
    # no comerse una palabra corta que casualmente sea todo hex (ej. "-decada").
    titulo = re.sub(r"^\d{4}-\d{2}-\d{2}-", "", ruta.stem)
    titulo = re.sub(r"-(?=[0-9a-f]*\d)[0-9a-f]{6,8}$", "", titulo)
    titulo = titulo.replace("-", " ").replace("_", " ").strip()
    ahora = datetime.datetime.now().astimezone().isoformat(timespec="seconds")
    cita = _origen_bundle(ruta, origen)

    fm = [
        "---",
        f"type: {tipo}",
        f"title: {titulo[:1].upper() + titulo[1:] if titulo else ruta.stem}",
        "description: <pendiente — una oración, al integrar>",
    ]
    if tipo == "Diagrama":
        fm += ['proyecto: <periodo-nombre o "transversal">', "clase: flujo", "version: " + ahora[:10]]
    else:
        fm += [f"formato: {ext.lstrip('.')}"]
    fm += [f"origen: {cita}", "tags: []", f"timestamp: {ahora}", "---", ""]

    doc = fm[:]
    doc += ["> Convertido automáticamente desde el original crudo por `a-markdown.py`. "
            "El original manda: ante cualquier duda, se cita y se revisa.", ""]
    if avisos:
        doc += ["## Avisos de conversión", ""] + [f"- {a}" for a in avisos] + [""]
    doc += [cuerpo, "", "# Citations", "", f"[1] Original: [{ruta.name}]({cita})"]
    return "\n".join(doc) + "\n", avisos


def main():
    import argparse

    ap = argparse.ArgumentParser(
        description="Convierte un insumo a markdown legible por un LLM (sin gastar tokens).")
    ap.add_argument("archivo", type=Path)
    ap.add_argument("--out", type=Path, help="directorio de salida (por defecto: junto al archivo)")
    ap.add_argument("--filas", type=int, default=FILAS_POR_DEFECTO,
                    help=f"filas por hoja en .xlsx (por defecto {FILAS_POR_DEFECTO})")
    ap.add_argument("--origen", help="puntero al raw a citar (ej. /raw/2026-07-16-propuesta-3f2a9c.pptx)")
    ap.add_argument("--stdout", action="store_true", help="imprimir el markdown en vez de escribirlo")
    args = ap.parse_args()

    try:
        doc, avisos = construir(args.archivo, args.filas, args.origen)
    except ErrorConversion as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    if args.stdout:
        print(doc, end="")
        return 0

    destino = (args.out or args.archivo.parent) / (args.archivo.stem + ".md")
    destino.parent.mkdir(parents=True, exist_ok=True)
    # newline="\n": sin esto Windows escribiría CRLF y el mismo insumo daría
    # diff según quién lo convirtió. El repo se comparte entre plataformas.
    destino.write_text(doc, encoding="utf-8", newline="\n")
    print(f"OK: {destino}  ({_miles(len(doc))} caracteres)")
    for a in avisos:
        print(f"  aviso: {a}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
